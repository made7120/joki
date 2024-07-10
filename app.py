import streamlit as st
import json
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from PyPDF2 import PdfReader
from openpyxl.utils.exceptions import InvalidFileException
import streamlit_authenticator as stauth
import bcrypt

# Definisikan kelas Buku, BukuDigital, BukuFisik, dan Perpustakaan
class Buku:
    def __init__(self, judul, penulis, tahun_terbit, status="tersedia"):
        self.judul = judul
        self.penulis = penulis
        self.tahun_terbit = tahun_terbit
        self.status = status

    def info_buku(self):
        return {
            "Judul": self.judul,
            "Penulis": self.penulis,
            "Tahun Terbit": str(self.tahun_terbit), 
            "Status": self.status
        }

class BukuDigital(Buku):
    def __init__(self, judul, penulis, tahun_terbit, ukuran_file, format_file, file_path, status="tersedia"):
        super().__init__(judul, penulis, tahun_terbit, status)
        self.ukuran_file = ukuran_file
        self.format_file = format_file
        self.file_path = file_path

    def info_buku(self):
        info = super().info_buku()
        info.update({"Ukuran File (MB)": self.ukuran_file, "Format File": self.format_file, "File Path": self.file_path})
        return info

class BukuFisik(Buku):
    def __init__(self, judul, penulis, tahun_terbit, jumlah_halaman, berat, status="tersedia"):
        super().__init__(judul, penulis, tahun_terbit, status)
        self.jumlah_halaman = jumlah_halaman
        self.berat = berat

    def info_buku(self):
        info = super().info_buku()
        info.update({"Jumlah Halaman": self.jumlah_halaman, "Berat (gram)": self.berat})
        return info

class Perpustakaan:
    def __init__(self, data_file="data_buku.json"):
        self.data_file = data_file
        self.daftar_buku = self.load_data()

    def tampilkan_semua_buku(self):
        return [buku.info_buku() for buku in self.daftar_buku]

    def cari_buku(self, judul):
        for buku in self.daftar_buku:
            if buku.judul.lower() == judul.lower():
                return buku
        return None

    def hapus_buku(self, judul):
        buku_dihapus = self.cari_buku(judul)
        if buku_dihapus:
            self.daftar_buku.remove(buku_dihapus)
            self.save_data()
            self.save_to_excel()
            st.success(f"Buku '{judul}' berhasil dihapus.")
        else:
            st.warning(f"Buku '{judul}' tidak ditemukan.")

    def pinjam_buku(self, judul):
        buku_dipinjam = self.cari_buku(judul)
        if buku_dipinjam:
            if buku_dipinjam.status == "tersedia":
                buku_dipinjam.status = "dipinjam"
                self.save_data()
                st.success(f"Buku '{judul}' berhasil dipinjam.")
            else:
                st.warning(f"Buku '{judul}' sedang tidak tersedia untuk dipinjam.")
        else:
            st.warning(f"Buku '{judul}' tidak ditemukan.")

    def kembalikan_buku(self, judul):
        buku_dikembalikan = self.cari_buku(judul)
        if buku_dikembalikan:
            if buku_dikembalikan.status == "dipinjam":
                buku_dikembalikan.status = "tersedia"
                self.save_data()
                st.success(f"Buku '{judul}' berhasil dikembalikan.")
            else:
                st.warning(f"Buku '{judul}' tidak sedang dipinjam.")
        else:
            st.warning(f"Buku '{judul}' tidak ditemukan.")

    def tampilkan_buku_dipinjam(self):
        return [buku.info_buku() for buku in self.daftar_buku if buku.status == "dipinjam"]

    def tampilkan_buku_dikembalikan(self):
        return [buku.info_buku() for buku in self.daftar_buku if buku.status == "tersedia"]

    def tambah_buku(self, buku):
        self.daftar_buku.append(buku)
        self.save_data()
        st.success(f"Buku '{buku.judul}' berhasil ditambahkan.")  
        self.save_to_excel()

    def load_data(self):
        if os.path.exists(self.data_file):
            try:
                dengan open(self.data_file, "r") sebagai file:
                    data = json.load(file)
                    daftar_buku = []
                    untuk item dalam data:
                        jika "Ukuran File (MB)" di item:
                            buku = BukuDigital(
                                judul=item["Judul"],
                                penulis=item["Penulis"],
                                tahun_terbit=item["Tahun Terbit"],
                                ukuran_file=item["Ukuran File (MB)"],
                                format_file=item["Format File"],
                                file_path=item.get("File Path", ""),
                                status=item["Status"]
                            )
                        elif "Jumlah Halaman" dalam item:
                            buku = BukuFisik(
                                judul=item["Judul"],
                                penulis=item["Penulis"],
                                tahun_terbit=item["Tahun Terbit"],
                                jumlah_halaman=item["Jumlah Halaman"],
                                berat=item["Berat (gram)"],
                                status=item["Status"]
                            )
                        lain:
                            buku = Buku(
                                judul=item["Judul"],
                                penulis=item["Penulis"],
                                tahun_terbit=item["Tahun Terbit"],
                                status=item["Status"]
                            )
                        daftar_buku.append(buku)
                    kembalikan daftar_buku
            kecuali json.JSONDecodeError:
                st.warning("File JSON kosong atau tidak valid. Menginisialisasi dengan daftar kosong.")
                kembalikan []
        kembalikan []

    def save_data(self):
        data = [buku.info_buku() untuk buku dalam self.daftar_buku]
        dengan open(self.data_file, "w") sebagai file:
            json.dump(data, file)

    def save_to_excel(self):
        coba:
            jika os.path.exists("daftar_buku.xlsx"):
                wb = load_workbook("daftar_buku.xlsx")
                ws = wb.active
            lain:
                wb = Workbook()
                ws = wb.active
                ws.append(["Judul", "Penulis", "Tahun Terbit", "Status"])

            # Hapus data yang ada
            untuk baris dalam ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                untuk sel di baris:
                    cell.value = None

            # Tambahkan data baru
            untuk buku dalam self.daftar_buku:
                ws.append([buku.judul, buku.penulis, buku.tahun_terbit, buku.status])
            wb.save("daftar_buku.xlsx")
        kecuali PermissionError:
            st.error("Gagal menyimpan ke 'daftar_buku.xlsx'. Pastikan file tidak sedang terbuka di program lain.")
        kecuali InvalidFileException:
            st.error("Format file tidak valid atau rusak.")
        kecuali Exception sebagai e:
            st.error(f"Gagal menyimpan ke 'daftar_buku.xlsx'. Kesalahan: {e}")

# Fungsi untuk membaca dan menampilkan isi PDF
def view_pdf(file_path):
    dengan open(file_path, "rb") sebagai f:
        pdf_reader = PdfReader(f)
        untuk page_num dalam range(len(pdf_reader.pages)):
            halaman = pdf_reader.pages[page_num]
            st.write(page.extract_text())

# Setup authenticator
usernames = ['username1', 'username2']
names = ['user1', 'user2']
passwords = ['123', '456']

hashed_passwords = [bcrypt.hashpw(p.encode('utf-8'), bcrypt.gensalt()).decode('utf-8') untuk p dalam passwords]

credentials = {
    "usernames": {
        usernames[0]: {"name": names[0], "password": hashed_passwords[0]},
        usernames[1]: {"name": names[1], "password": hashed_passwords[1]}
    }
}

authenticator = stauth.Authenticate(credentials, 'cookie_name', 'signature_key', cookie_expiry_days=30)

# Fungsi untuk menampilkan antarmuka Streamlit
def main():
    st.title("Baca Buku Digital")

    # Contoh path file PDF dari buku digital
    file_path = "path_to_your_pdf_file.pdf"  # Ganti dengan path file PDF Anda

    jika st.button("Baca Buku"):
        jika os.path.exists(file_path):
            view_pdf(file_path)
        lain:
            st.error(f"File PDF '{file_path}' tidak ditemukan.")

# Inisialisasi perpustakaan
perpustakaan = Perpustakaan()

# Login dan Register
nama, authentication_status, username = authenticator.login('Login', 'main')

jika authentication_status:
    st.sidebar.title(f"Selamat Datang {name}")
    menu = [
        {"label": "Tambah Buku Digital", "icon": "💻️"},
        {"label": "Tambah Buku Fisik", "icon": "📚️"},
        {"label": "Tampilkan Semua Buku", "icon": "📖️"},
        {"label": "Edit Buku", "icon": "✏️"},
        {"label": "Hapus Buku", "icon": "❌"},
        {"label": "Pinjam Buku", "icon": "📝️"},
        {"label": "Kembalikan Buku", "icon": "📚️‍♂️"},
        {"label": "Tampilkan Buku yang Dipinjam", "icon": "📊️"},
        {"label": "Tampilkan Buku yang Dikembalikan", "icon": "📈️"}
    ]

    # Fungsi untuk menampilkan antarmuka Streamlit
    def library_app():
        st.markdown(
            """
            <style>
            .stApp {
                background-image: url("https://ae01.alicdn.com/kf/H2ff989d930434f33b80e36fdb0dd11e8c/Latar-Belakang-Rak-Buku-Latar-Belakang-Rak-Buku-Latar-Belakang-Kantor-Perpustakaan-untuk-Konferensi-Video-Buku.jpg");
                background-size: cover;
            }
            /* CSS untuk memindahkan sidebar ke kiri */
            .st-emotion-cache-6qob1r{
                background-color: #66CDAA;
            }
            .css-1d391kg {
                left: auto;
                right: 0;
            }
            .css-1v0mbdj.e1fvgfdy4 {
                color: #A52A2A;  /* Warna teks */
                font-size: 50px;  /* Ukuran teks */
            }
            .css-14xtw13.e8zbici2 {
                color: #A52A2A;  /* Warna teks */
                font-size: 40px;  /* Ukuran teks */
            }
            </style>
            """,
            unsafe_allow_html=True
        )
        st.sidebar.title("MENU PERPUSTAKAAN DIGITAL")

        menu_labels = [f"{item['icon']} {item['label']}" untuk item dalam menu]
        choice = st.sidebar.selectbox("Pilih Menu:", menu_labels)

        st.title("Selamat Datang di Perpustakaan Digital")

        jika choice == "💻️ Tambah Buku Digital":
            st.subheader("Tambah Buku Digital")
            st.image("digital_book.jpg", width=200)
            judul = st.text_input("Judul", label_visibility="visible")
            penulis = st.text_input("Penulis", label_visibility="visible")
            tahun_terbit = st.number_input("Tahun Terbit", min_value=0, max_value=2024, step=1, label_visibility="visible")
            ukuran_file = st.number_input("Ukuran File (MB)", min_value=0.0, step=0.1, label_visibility="visible")
            format_file = st.selectbox("Format File", ["PDF", "EPUB", "MOBI"], label_visibility="visible")
            uploaded_file = st.file_uploader("Unggah File Buku", type=["pdf", "epub", "mobi"])

            jika st.button("Tambah Buku Digital"):
                jika uploaded_file is not None:
                    jika tidak os.path.exists("uploads"):
                        os.makedirs("uploads")
                    file_path = os.path.join("uploads", uploaded_file.name)
                    dengan open(file_path, "wb") sebagai f:
                        f.write(uploaded_file.getbuffer())
                    buku_digital = BukuDigital(judul, penulis, tahun_terbit, ukuran_file, format_file, file_path)
                    perpustakaan.tambah_buku(buku_digital)
                lain:
                    st.error("Harap unggah file buku terlebih dahulu.")

        elif choice == "📚️ Tambah Buku Fisik":
            st.subheader("Tambah Buku Fisik")
            st.image("physical_book.jpg", width=200)
            judul = st.text_input("Judul", label_visibility="visible")
            penulis = st.text_input("Penulis", label_visibility="visible")
            tahun_terbit = st.number_input("Tahun Terbit", min_value=0, max_value=2024, step=1, label_visibility="visible")
            jumlah_halaman = st.number_input("Jumlah Halaman", min_value=1, step=1, label_visibility="visible")
            berat = st.number_input("Berat (gram)", min_value=1.0, step=0.1, label_visibility="visible")

            jika st.button("Tambah Buku Fisik"):
                buku_fisik = BukuFisik(judul, penulis, tahun_terbit, jumlah_halaman, berat)
                perpustakaan.tambah_buku(buku_fisik)

        elif choice == "📖️ Tampilkan Semua Buku":
            st.subheader("Daftar Semua Buku")
            daftar_buku = perpustakaan.tampilkan_semua_buku()
            jika daftar_buku:
                df = pd.DataFrame(daftar_buku)
                st.dataframe(df)

                untuk buku dalam daftar_buku:
                    jika "File Path" di buku:
                        jika buku["File Path"] dan os.path.exists(buku["File Path"]):
                            jika st.button(f"Baca Buku: {buku['Judul']}"):
                                dengan open(buku["File Path"], "rb") sebagai f:
                                    st.download_button(label="Klik untuk Membaca", data=f, file_name=buku["Judul"], mime="application/octet-stream")
                        lain:
                            st.error(f"File untuk buku '{buku['Judul']}' tidak ditemukan.")
            lain:
                st.warning("Tidak ada buku yang tersedia.")

        elif choice == "✏️ Edit Buku":
            st.subheader("Edit Buku")
            daftar_buku = perpustakaan.tampilkan_semua_buku()
            jika daftar_buku:
                df = pd.DataFrame(daftar_buku)
                st.dataframe(df)

                judul_buku_edit = st.text_input("Masukkan judul buku yang akan diedit:")
                buku_diedit = perpustakaan.cari_buku(judul_buku_edit)

                jika buku_diedit:
                    new_judul = st.text_input("Judul Baru", buku_diedit.judul)
                    new_penulis = st.text_input("Penulis Baru", buku_diedit.penulis)
                    new_tahun_terbit = st.number_input("Tahun Terbit Baru", min_value=0, max_value=2024, step=1, value=int(buku_diedit.tahun_terbit))
                    jika isinstance(buku_diedit, BukuDigital):
                        new_ukuran_file = st.number_input("Ukuran File (MB) Baru", min_value=0.0, step=0.1, value=float(buku_diedit.ukuran_file))
                        new_format_file = st.selectbox("Format File Baru", ["PDF", "EPUB", "MOBI"], index=["PDF", "EPUB", "MOBI"].index(buku_diedit.format_file))
                        uploaded_file = st.file_uploader("Unggah File Buku Baru", type=["pdf", "epub", "mobi"])
                        jika st.button("Simpan Perubahan"):
                            buku_diedit.judul = new_judul
                            buku_diedit.penulis = new_penulis
                            buku_diedit.tahun_terbit = new_tahun_terbit
                            buku_diedit.ukuran_file = new_ukuran_file
                            buku_diedit.format_file = new_format_file
                            jika uploaded_file is not None:
                                file_path = os.path.join("uploads", uploaded_file.name)
                                dengan open(file_path, "wb") sebagai f:
                                    f.write(uploaded_file.getbuffer())
                                buku_diedit.file_path = file_path
                            perpustakaan.save_data()
                            st.success("Buku berhasil diperbarui.")
                    elif isinstance(buku_diedit, BukuFisik):
                        new_jumlah_halaman = st.number_input("Jumlah Halaman Baru", min_value=1, step=1, value=int(buku_diedit.jumlah_halaman))
                        new_berat = st.number_input("Berat (gram) Baru", min_value=1.0, step=0.1, value=float(buku_diedit.berat))
                        jika st.button("Simpan Perubahan"):
                            buku_diedit.judul = new_judul
                            buku_diedit.penulis = new_penulis
                            buku_diedit.tahun_terbit = new_tahun_terbit
                            buku_diedit.jumlah_halaman = new_jumlah_halaman
                            buku_diedit.berat = new_berat
                            perpustakaan.save_data()
                            st.success("Buku berhasil diperbarui.")
                    lain:
                        jika st.button("Simpan Perubahan"):
                            buku_diedit.judul = new_judul
                            buku_diedit.penulis = new_penulis
                            buku_diedit.tahun_terbit = new_tahun_terbit
                            perpustakaan.save_data()
                            st.success("Buku berhasil diperbarui.")
                lain:
                    st.warning("Buku tidak ditemukan.")
            lain:
                st.warning("Tidak ada buku yang tersedia.")

        elif choice == "❌ Hapus Buku":
            st.subheader("Hapus Buku")
            daftar_buku = perpustakaan.tampilkan_semua_buku()
            jika daftar_buku:
                df = pd.DataFrame(daftar_buku)
                st.dataframe(df)

                judul_buku_hapus = st.text_input("Masukkan judul buku yang akan dihapus:")
                jika st.button("Hapus Buku"):
                    perpustakaan.hapus_buku(judul_buku_hapus)
            lain:
                st.warning("Tidak ada buku yang tersedia.")

        elif choice == "📝️ Pinjam Buku":
            st.subheader("Pinjam Buku")
            daftar_buku = perpustakaan.tampilkan_buku_dikembalikan()
            jika daftar_buku:
                df = pd.DataFrame(daftar_buku)
                st.dataframe(df)

                judul_buku_pinjam = st.text_input("Masukkan judul buku yang akan dipinjam:")
                jika st.button("Pinjam Buku"):
                    hasil = perpustakaan.pinjam_buku(judul_buku_pinjam)
                    st.success(hasil)
            lain:
                st.warning("Tidak ada buku yang tersedia untuk dipinjam.")

        elif choice == "📚️‍♂️ Kembalikan Buku":
            st.subheader("Kembalikan Buku")
            daftar_buku = perpustakaan.tampilkan_buku_dipinjam()
            jika daftar_buku:
                df = pd.DataFrame(daftar_buku)
                st.dataframe(df)

                judul_buku_kembalikan = st.text_input("Masukkan judul buku yang akan dikembalikan:")
                jika st.button("Kembalikan Buku"):
                    hasil = perpustakaan.kembalikan_buku(judul_buku_kembalikan)
                    st.success(hasil)
            lain:
                st.warning("Tidak ada buku yang sedang dipinjam.")

        elif choice == "📊️ Tampilkan Buku yang Dipinjam":
            st.subheader("Daftar Buku yang Dipinjam")
            daftar_buku = perpustakaan.tampilkan_buku_dipinjam()
            jika daftar_buku:
                df = pd.DataFrame(daftar_buku)
                st.dataframe(df)
            lain:
                st.warning("Tidak ada buku yang sedang dipinjam.")

        elif choice == "📈️ Tampilkan Buku yang Dikembalikan":
            st.subheader("Daftar Buku yang Dikembalikan")
            daftar_buku = perpustakaan.tampilkan_buku_dikembalikan()
            jika daftar_buku:
                df = pd.DataFrame(daftar_buku)
                st.dataframe(df)
            lain:
                st.warning("Tidak ada buku yang tersedia untuk dipinjam atau telah dikembalikan.")

    library_app()
lain:
    st.warning('Harap login untuk mengakses sistem perpustakaan.')
    st.stop()

jika authentication_status == False:
    st.error('Username/password salah')

jika authentication_status == None:
    st.warning('Harap masukkan username dan password')

jika st.button('Logout'):
    authenticator.logout('main', 'sidebar')

jika __name__ == "__main__":
    main()
